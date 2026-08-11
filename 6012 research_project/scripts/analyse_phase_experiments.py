"""Batch analyse the 21 randomised three-marker phase experiments.

Expected filename format (case-insensitive):
    run01_phase270_rep2.mp4

On its first run the script creates/updates ``experiment_manifest.csv`` from
the videos in EXPERIMENT_VIDEO_DIR. It asks for one tight rectangle around the
red LED and remembers that ROI for every video. Results are written to one Excel
workbook, per-run CSV files, diagnostic plots and optional annotated videos.
"""

from __future__ import annotations

import argparse
import itertools
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path

import cv2
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd


# =============================================================================
# USER SETTINGS
# =============================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
EXPERIMENT_VIDEO_DIR = SCRIPT_DIR / "doe_phase_experiment_videos"
CALIBRATION_NPZ = (
    SCRIPT_DIR / "doe_session_calibration_outputs" / "doe_session_calibration.npz"
)
MANIFEST_CSV = EXPERIMENT_VIDEO_DIR / "experiment_manifest.csv"
TRACKING_SETTINGS_JSON = EXPERIMENT_VIDEO_DIR / "tracking_settings.json"
OUTPUT_DIR = SCRIPT_DIR / "doe_phase_analysis_outputs"
WORKBOOK_PATH = OUTPUT_DIR / "phase_experiment_results.xlsx"

VIDEO_EXTENSIONS = {".mp4", ".mov", ".m4v", ".avi"}
EXPECTED_PHASES = [72, 90, 120, 180, 240, 270, 288]
EXPECTED_REPEATS = 3

# Fixed experimental conditions recorded in the workbook.
WAVEFORM = "Rounded triangle"
DRIVE_LENGTH_MM = 70.0
SPEED_MULTIPLIER = 1.0
CILIA_SPACING_MM = 87.0
SURFACE_MATERIAL = "High-friction surface"

WRITE_ANNOTATED_VIDEOS = True
DISPLAY_PROCESSING = False

# LED detection. Select a tight ROI around only the LED when prompted.
LED_BASELINE_SECONDS = 1.0
LED_MIN_ON_FRAMES = 2
LED_MIN_GAP_SECONDS = 0.25
LED_THRESHOLD_MIN_RISE = 10.0

# The phase experiments used a fixed 3 s gait cycle. An LED interval outside
# this tolerance is a partial start/stop interval, not a complete gait cycle.
EXPECTED_CYCLE_DURATION_S = 3.0
COMPLETE_CYCLE_DURATION_TOLERANCE = 0.20

# Green marker detection. Diameter and HSV are loaded from the calibration.
MIN_DIAMETER_FACTOR = 0.58
MAX_DIAMETER_FACTOR = 1.55
MIN_CIRCULARITY = 0.48
MAX_TRACKING_JUMP_PX = 115.0
MAX_INTERPOLATION_GAP_FRAMES = 5

# Metrics use the smoothed centroid. Small frame-to-frame changes below this
# value are treated as camera/detection noise rather than physical slip.
SMOOTHING_WINDOW_FRAMES = 5
MOVEMENT_NOISE_FLOOR_MM = 0.20


RUN_SUMMARY_COLUMNS = [
    "Randomised Run Order", "Test ID", "Video Filename",
    "Phase Shift Command (deg)", "Equivalent Signed Phase (deg)",
    "Repeat Number", "Start Blink Time (s)", "Finish Crossing Time (s)",
    "Transport Time (s)", "Net Travel Distance (mm)",
    "Mean Transport Speed (mm/s)", "Median Transport Speed (mm/s)",
    "Number of Complete Cycles", "Mean Net Advance per Cycle (mm)",
    "SD Net Advance per Cycle (mm)", "Total Backward Travel (mm)",
    "Backward Travel Percentage (%)", "Maximum Backward Slip in One Cycle (mm)",
    "Vertical Centroid Peak-to-Peak (mm)", "Vertical Centroid RMS (mm)",
    "Mean Pitch Angle (deg)", "Pitch Peak-to-Peak (deg)",
    "Maximum Absolute Pitch Angle (deg)", "Green Marker Detection Rate (%)",
    "Frames Requiring Interpolation", "Finish Crossing Detected",
    "Quality-Control Status", "Notes",
]

CYCLE_COLUMNS = [
    "Test ID", "Cycle Number", "Cycle Start Time (s)", "Cycle End Time (s)",
    "Measured Cycle Duration (s)", "Centroid Start X (mm)",
    "Centroid End X (mm)", "Net Advance (mm)", "Total Forward Movement (mm)",
    "Total Backward Movement (mm)", "Maximum Backward Excursion (mm)",
    "Vertical Peak-to-Peak (mm)", "Mean Pitch Angle (deg)",
    "Pitch Peak-to-Peak (deg)",
]

REJECTED_CYCLE_COLUMNS = [
    "Test ID", "Candidate Interval Number", "Interval Start Time (s)",
    "Interval End Time (s)", "Measured Interval Duration (s)",
    "Expected Cycle Duration (s)", "Rejection Reason",
]

FRAME_COLUMNS = [
    "Test ID", "Frame Number", "Video Time (s)",
    "Elapsed Experiment Time (s)", "LED Detected", "Marker 1 X (mm)",
    "Marker 1 Y (mm)", "Marker 2 X (mm)", "Marker 2 Y (mm)",
    "Marker 3 X (mm)", "Marker 3 Y (mm)", "Object Centroid X (mm)",
    "Object Centroid Y (mm)", "Pitch Angle (deg)",
    "Forward Velocity (mm/s)", "Number of Markers Detected",
    "Interpolated Frame", "Finish Line Crossed",
]


@dataclass
class Calibration:
    camera_matrix: np.ndarray
    distortion: np.ndarray
    new_camera_matrix: np.ndarray
    image_size: tuple[int, int]
    homography: np.ndarray
    finish_line: np.ndarray
    finish_points: np.ndarray
    hsv_lower: np.ndarray
    hsv_upper: np.ndarray
    dot_diameter_mm: float
    travel_direction: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--reset-led-roi", action="store_true",
        help="Forget the saved LED rectangle and select it again.",
    )
    parser.add_argument(
        "--no-overlay", action="store_true",
        help="Skip annotated MP4 creation to reduce processing time/storage.",
    )
    return parser.parse_args()


def load_calibration(path: Path) -> Calibration:
    if not path.exists():
        raise FileNotFoundError(f"Session calibration not found: {path}")
    with np.load(path) as data:
        return Calibration(
            camera_matrix=np.asarray(data["camera_matrix"], float),
            distortion=np.asarray(data["dist_coeffs"], float),
            new_camera_matrix=np.asarray(data["new_camera_matrix"], float),
            image_size=tuple(int(x) for x in data["image_size"]),
            homography=np.asarray(data["image_to_plane_mm"], float),
            finish_line=np.asarray(data["finish_line_abc_px"], float),
            finish_points=np.asarray(data["finish_line_points_px"], float),
            hsv_lower=np.asarray(data["green_hsv_lower"], np.uint8),
            hsv_upper=np.asarray(data["green_hsv_upper"], np.uint8),
            dot_diameter_mm=float(data["expected_green_dot_diameter_mm"]),
            travel_direction=str(data["travel_direction"]),
        )


def signed_phase(phase: int) -> int:
    return phase if phase <= 180 else phase - 360


def filename_fields(path: Path) -> tuple[float, float, float]:
    text = path.stem
    run_match = re.search(r"run[ _-]?(\d+)", text, re.IGNORECASE)
    phase_match = re.search(r"phase[ _-]?(\d+)", text, re.IGNORECASE)
    repeat_match = re.search(r"rep(?:eat)?[ _-]?(\d+)", text, re.IGNORECASE)
    return (
        float(run_match.group(1)) if run_match else math.nan,
        float(phase_match.group(1)) if phase_match else math.nan,
        float(repeat_match.group(1)) if repeat_match else math.nan,
    )


def discover_videos() -> list[Path]:
    EXPERIMENT_VIDEO_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(
        path for path in EXPERIMENT_VIDEO_DIR.iterdir()
        if path.is_file() and path.suffix.lower() in VIDEO_EXTENSIONS
        and "overlay" not in path.stem.lower()
    )


def create_or_update_manifest(videos: list[Path]) -> pd.DataFrame:
    columns = [
        "Include", "Randomised Run Order", "Test ID", "Video Filename",
        "Phase Shift (deg)", "Repeat Number", "Waveform", "Drive Length (mm)",
        "Speed Multiplier", "Cilia Spacing (mm)", "Surface Material",
        "Recording Date", "Operator Notes",
    ]
    existing = pd.read_csv(MANIFEST_CSV) if MANIFEST_CSV.exists() else pd.DataFrame()
    existing_by_name = (
        existing.set_index("Video Filename").to_dict("index")
        if not existing.empty and "Video Filename" in existing else {}
    )
    rows: list[dict[str, object]] = []
    for video in videos:
        run, phase, repeat = filename_fields(video)
        old = existing_by_name.get(video.name, {})
        phase_value = old.get("Phase Shift (deg)", phase)
        repeat_value = old.get("Repeat Number", repeat)
        test_id = old.get("Test ID", "")
        if not test_id and not pd.isna(phase_value) and not pd.isna(repeat_value):
            test_id = f"phase{int(phase_value):03d}_rep{int(repeat_value)}"
        rows.append({
            "Include": old.get("Include", 1),
            "Randomised Run Order": old.get("Randomised Run Order", run),
            "Test ID": test_id,
            "Video Filename": video.name,
            "Phase Shift (deg)": phase_value,
            "Repeat Number": repeat_value,
            "Waveform": old.get("Waveform", WAVEFORM),
            "Drive Length (mm)": old.get("Drive Length (mm)", DRIVE_LENGTH_MM),
            "Speed Multiplier": old.get("Speed Multiplier", SPEED_MULTIPLIER),
            "Cilia Spacing (mm)": old.get("Cilia Spacing (mm)", CILIA_SPACING_MM),
            "Surface Material": old.get("Surface Material", SURFACE_MATERIAL),
            "Recording Date": old.get("Recording Date", ""),
            "Operator Notes": old.get("Operator Notes", ""),
        })
    manifest = pd.DataFrame(rows, columns=columns)
    if not manifest.empty:
        manifest = manifest.sort_values(
            "Randomised Run Order", na_position="last"
        ).reset_index(drop=True)
    manifest.to_csv(MANIFEST_CSV, index=False)
    return manifest


def validate_manifest(manifest: pd.DataFrame) -> None:
    if manifest.empty:
        raise SystemExit(
            f"No experiment videos found. Put them in:\n{EXPERIMENT_VIDEO_DIR}"
        )
    included = manifest[pd.to_numeric(manifest["Include"], errors="coerce").fillna(0) != 0]
    required = ["Randomised Run Order", "Test ID", "Phase Shift (deg)", "Repeat Number"]
    missing = [name for name in required if included[name].isna().any()]
    blank_test_ids = included["Test ID"].astype(str).str.strip().eq("").any()
    if missing or blank_test_ids:
        raise SystemExit(
            f"Complete the missing run/phase/repeat fields in:\n{MANIFEST_CSV}\n"
            "The recommended filename is run01_phase270_rep2.mp4."
        )
    phases = set(pd.to_numeric(included["Phase Shift (deg)"]).astype(int))
    unexpected = phases - set(EXPECTED_PHASES)
    if unexpected:
        print(f"WARNING: unexpected phase values in manifest: {sorted(unexpected)}")
    if len(included) != len(EXPECTED_PHASES) * EXPECTED_REPEATS:
        print(f"WARNING: expected 21 included videos, found {len(included)}.")


def undistort(frame: np.ndarray, calibration: Calibration) -> np.ndarray:
    return cv2.undistort(
        frame, calibration.camera_matrix, calibration.distortion,
        None, calibration.new_camera_matrix
    )


def select_or_load_led_roi(
    first_video: Path, calibration: Calibration, reset: bool
) -> tuple[int, int, int, int]:
    if TRACKING_SETTINGS_JSON.exists() and not reset:
        settings = json.loads(TRACKING_SETTINGS_JSON.read_text(encoding="utf-8"))
        roi = settings.get("led_roi_px")
        if roi and len(roi) == 4:
            return tuple(int(x) for x in roi)

    capture = cv2.VideoCapture(str(first_video))
    if not capture.isOpened():
        raise RuntimeError(f"Could not open {first_video}")
    fps = float(capture.get(cv2.CAP_PROP_FPS))
    capture.set(cv2.CAP_PROP_POS_FRAMES, max(0, int(round(0.5 * fps))))
    ok, frame = capture.read()
    capture.release()
    if not ok:
        raise RuntimeError("Could not read the first video for LED selection.")
    frame = undistort(frame, calibration)
    print("\nDrag a TIGHT rectangle around the red LED, then press Enter/Space.")
    print("Press C to cancel the current rectangle and draw it again.")
    roi = cv2.selectROI("Select red LED only", frame, showCrosshair=True, fromCenter=False)
    cv2.destroyWindow("Select red LED only")
    x, y, width, height = (int(value) for value in roi)
    if width < 3 or height < 3:
        raise SystemExit("LED ROI selection was cancelled or too small.")
    TRACKING_SETTINGS_JSON.write_text(
        json.dumps({"led_roi_px": [x, y, width, height]}, indent=2),
        encoding="utf-8",
    )
    return x, y, width, height


def image_to_mm(points_px: np.ndarray, homography: np.ndarray) -> np.ndarray:
    return cv2.perspectiveTransform(
        np.asarray(points_px, np.float32).reshape(-1, 1, 2), homography
    ).reshape(-1, 2)


def local_mm_per_pixel(point_px: np.ndarray, homography: np.ndarray) -> float:
    x, y = (float(v) for v in point_px)
    mapped = image_to_mm(np.asarray([[x, y], [x + 1, y], [x, y + 1]]), homography)
    return math.sqrt(
        float(np.linalg.norm(mapped[1] - mapped[0]))
        * float(np.linalg.norm(mapped[2] - mapped[0]))
    )


def led_score(frame: np.ndarray, roi: tuple[int, int, int, int]) -> float:
    x, y, width, height = roi
    patch = frame[y:y + height, x:x + width]
    if patch.size == 0:
        return math.nan
    blue, green, red = cv2.split(patch.astype(np.float32))
    red_excess = red - 0.55 * green - 0.45 * blue
    # A tight ROI makes the brightest 10% a stable measure of LED illumination.
    cutoff = np.percentile(red_excess, 90)
    brightest = red_excess[red_excess >= cutoff]
    return float(np.mean(brightest)) if brightest.size else float(cutoff)


def marker_candidates(frame: np.ndarray, calibration: Calibration) -> list[dict[str, float]]:
    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
    mask = cv2.inRange(hsv, calibration.hsv_lower, calibration.hsv_upper)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    candidates: list[dict[str, float]] = []
    for contour in contours:
        area = float(cv2.contourArea(contour))
        perimeter = float(cv2.arcLength(contour, True))
        moments = cv2.moments(contour)
        if area < 30 or perimeter <= 0 or moments["m00"] == 0:
            continue
        point = np.asarray(
            [moments["m10"] / moments["m00"], moments["m01"] / moments["m00"]],
            np.float32,
        )
        diameter_px = math.sqrt(4 * area / math.pi)
        diameter_mm = diameter_px * local_mm_per_pixel(point, calibration.homography)
        circularity = 4 * math.pi * area / perimeter**2
        if not (
            calibration.dot_diameter_mm * MIN_DIAMETER_FACTOR
            <= diameter_mm
            <= calibration.dot_diameter_mm * MAX_DIAMETER_FACTOR
            and circularity >= MIN_CIRCULARITY
        ):
            continue
        candidates.append({
            "x_px": float(point[0]), "y_px": float(point[1]),
            "diameter_mm": diameter_mm, "circularity": circularity,
            "quality": abs(diameter_mm - calibration.dot_diameter_mm)
                       + 8 * max(0.0, 0.8 - circularity),
        })
    return sorted(candidates, key=lambda item: item["quality"])[:8]


def assign_markers(
    candidates: list[dict[str, float]],
    previous: np.ndarray | None,
    velocity: np.ndarray | None,
) -> tuple[np.ndarray, int]:
    result = np.full((3, 2), np.nan, float)
    if not candidates:
        return result, 0
    points = np.asarray([[x["x_px"], x["y_px"]] for x in candidates], float)
    if previous is None or np.count_nonzero(np.isfinite(previous[:, 0])) < 2:
        chosen = sorted(candidates[:3], key=lambda item: item["x_px"])
        for index, item in enumerate(chosen):
            result[index] = [item["x_px"], item["y_px"]]
        return result, len(chosen)

    prediction = previous.copy()
    if velocity is not None:
        valid_velocity = np.isfinite(velocity).all(axis=1)
        prediction[valid_velocity] += velocity[valid_velocity]
    valid_ids = [index for index in range(3) if np.isfinite(prediction[index]).all()]
    used: set[int] = set()
    # For three available candidates, solve the complete identity assignment.
    if len(candidates) >= 3 and len(valid_ids) == 3:
        best: tuple[float, tuple[int, int, int]] | None = None
        for permutation in itertools.permutations(range(len(candidates)), 3):
            distances = [
                np.linalg.norm(points[permutation[i]] - prediction[i]) for i in range(3)
            ]
            if max(distances) > MAX_TRACKING_JUMP_PX:
                continue
            quality = sum(candidates[j]["quality"] for j in permutation)
            cost = float(sum(distances) + 0.15 * quality)
            if best is None or cost < best[0]:
                best = (cost, permutation)
        if best is not None:
            for marker_id, candidate_id in enumerate(best[1]):
                result[marker_id] = points[candidate_id]
            return result, 3

    # Gracefully retain identities when one or two dots are momentarily hidden.
    pairs: list[tuple[float, int, int]] = []
    for marker_id in valid_ids:
        for candidate_id, point in enumerate(points):
            pairs.append((float(np.linalg.norm(point - prediction[marker_id])),
                          marker_id, candidate_id))
    for distance, marker_id, candidate_id in sorted(pairs):
        if distance > MAX_TRACKING_JUMP_PX or candidate_id in used:
            continue
        if np.isfinite(result[marker_id]).all():
            continue
        result[marker_id] = points[candidate_id]
        used.add(candidate_id)
    return result, int(np.count_nonzero(np.isfinite(result[:, 0])))


def rising_edges(on: np.ndarray, minimum_gap_frames: int) -> list[int]:
    indices: list[int] = []
    starts = np.flatnonzero(on & ~np.r_[False, on[:-1]])
    for index in starts:
        if not indices or index - indices[-1] >= minimum_gap_frames:
            indices.append(int(index))
    return indices


def interpolate_short_gaps(series: pd.Series, limit: int) -> tuple[pd.Series, pd.Series]:
    original_missing = series.isna()
    filled = series.interpolate(method="linear", limit=limit, limit_area="inside")
    return filled, original_missing & filled.notna()


def pitch_from_markers(frame: pd.DataFrame) -> pd.Series:
    # Plane Y increases image-down, hence the minus sign gives conventional Y-up.
    dx = frame["Marker 3 X (mm)"] - frame["Marker 1 X (mm)"]
    dy_up = -(frame["Marker 3 Y (mm)"] - frame["Marker 1 Y (mm)"])
    return pd.Series(np.degrees(np.arctan2(dy_up, dx)), index=frame.index)


def first_finish_crossing(
    marker_px: np.ndarray, start_index: int, finish_line: np.ndarray
) -> tuple[int | None, float]:
    values = marker_px @ finish_line[:2] + finish_line[2]
    valid_start = next(
        (i for i in range(start_index, len(values)) if np.isfinite(values[i])), None
    )
    if valid_start is None:
        return None, math.nan
    initial_sign = 1.0 if values[valid_start] >= 0 else -1.0
    previous = valid_start
    for index in range(valid_start + 1, len(values)):
        if not np.isfinite(values[index]):
            continue
        if values[index] * initial_sign <= 0:
            denominator = abs(values[previous]) + abs(values[index])
            fraction = abs(values[previous]) / denominator if denominator else 0.0
            return index, float(previous + fraction * (index - previous))
        previous = index
    return None, math.nan


def track_raw_video(
    video_path: Path,
    calibration: Calibration,
    led_roi: tuple[int, int, int, int],
) -> tuple[pd.DataFrame, dict[str, float | int], np.ndarray]:
    capture = cv2.VideoCapture(str(video_path))
    if not capture.isOpened():
        raise RuntimeError(f"Could not open video: {video_path}")
    fps = float(capture.get(cv2.CAP_PROP_FPS))
    count = int(capture.get(cv2.CAP_PROP_FRAME_COUNT))
    width = int(capture.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(capture.get(cv2.CAP_PROP_FRAME_HEIGHT))
    if (width, height) != calibration.image_size:
        capture.release()
        raise RuntimeError(
            f"Video is {width}x{height}, but calibration is "
            f"{calibration.image_size[0]}x{calibration.image_size[1]}. "
            "Do not crop, zoom or change recording resolution."
        )

    records: list[dict[str, float | int]] = []
    marker_px_history: list[np.ndarray] = []
    previous: np.ndarray | None = None
    previous_previous: np.ndarray | None = None
    for frame_index in range(count):
        ok, raw = capture.read()
        if not ok:
            break
        frame = undistort(raw, calibration)
        candidates = marker_candidates(frame, calibration)
        velocity = None
        if previous is not None and previous_previous is not None:
            velocity = previous - previous_previous
        assigned, detected = assign_markers(candidates, previous, velocity)
        # Keep the last known positions for matching, without claiming a measured
        # point in the output. Short gaps are interpolated only after the pass.
        if previous is None:
            matching_previous = assigned.copy()
        else:
            matching_previous = previous.copy()
            valid = np.isfinite(assigned).all(axis=1)
            matching_previous[valid] = assigned[valid]
        previous_previous = previous.copy() if previous is not None else None
        previous = matching_previous
        marker_px_history.append(assigned)
        records.append({
            "Frame Number": frame_index,
            "Video Time (s)": frame_index / fps,
            "LED Score": led_score(frame, led_roi),
            "Number of Markers Detected": detected,
        })
        if DISPLAY_PROCESSING:
            display = frame.copy()
            for marker_id, point in enumerate(assigned):
                if np.isfinite(point).all():
                    cv2.circle(display, tuple(np.round(point).astype(int)), 12,
                               (0, 255, 0), 3, cv2.LINE_AA)
                    cv2.putText(display, str(marker_id + 1),
                                tuple(np.round(point + [14, -10]).astype(int)),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.imshow("Processing (Q to stop)", cv2.resize(display, (960, 540)))
            if cv2.waitKey(1) & 0xFF in (ord("q"), 27):
                capture.release()
                cv2.destroyAllWindows()
                raise KeyboardInterrupt("Processing cancelled.")
    capture.release()
    cv2.destroyAllWindows()

    data = pd.DataFrame(records)
    marker_px = np.asarray(marker_px_history, float)
    actual_count = len(data)
    if actual_count == 0:
        raise RuntimeError("No frames were decoded.")
    metadata = {
        "fps": fps, "frame_count": actual_count, "width": width, "height": height
    }
    return data, metadata, marker_px


def detect_led_blinks(data: pd.DataFrame, fps: float) -> tuple[np.ndarray, list[int], float]:
    score = data["LED Score"].to_numpy(float)
    smooth = pd.Series(score).rolling(3, center=True, min_periods=1).median().to_numpy()
    baseline_count = max(3, min(len(smooth), int(round(LED_BASELINE_SECONDS * fps))))
    baseline = smooth[:baseline_count]
    median = float(np.nanmedian(baseline))
    mad = float(np.nanmedian(np.abs(baseline - median)))
    threshold = median + max(LED_THRESHOLD_MIN_RISE, 7.0 * 1.4826 * mad)
    raw_on = smooth >= threshold
    # Reject isolated red reflections; a 150 ms pulse is normally 4-5 frames.
    on = np.zeros_like(raw_on, dtype=bool)
    changes = np.flatnonzero(np.diff(np.r_[False, raw_on, False]))
    for start, end in changes.reshape(-1, 2):
        if end - start >= LED_MIN_ON_FRAMES:
            on[start:end] = True
    edges = rising_edges(on, max(1, int(round(LED_MIN_GAP_SECONDS * fps))))
    data["LED Score Smoothed"] = smooth
    data["LED Detected"] = on
    return on, edges, threshold


def prepare_frame_data(
    test_id: str,
    raw: pd.DataFrame,
    marker_px: np.ndarray,
    calibration: Calibration,
    fps: float,
    start_frame: int,
    finish_frame: int | None,
    crossing_frame_float: float,
) -> tuple[pd.DataFrame, np.ndarray]:
    count = len(raw)
    interpolated_any = np.zeros(count, dtype=bool)
    marker_px_filled = marker_px.copy()
    for marker in range(3):
        for axis in range(2):
            series = pd.Series(marker_px[:, marker, axis])
            filled, was_interpolated = interpolate_short_gaps(
                series, MAX_INTERPOLATION_GAP_FRAMES
            )
            marker_px_filled[:, marker, axis] = filled.to_numpy()
            interpolated_any |= was_interpolated.to_numpy()

    marker_mm = np.full_like(marker_px_filled, np.nan, dtype=float)
    for marker in range(3):
        valid = np.isfinite(marker_px_filled[:, marker]).all(axis=1)
        if np.any(valid):
            marker_mm[valid, marker] = image_to_mm(
                marker_px_filled[valid, marker], calibration.homography
            )

    output = pd.DataFrame({
        "Test ID": test_id,
        "Frame Number": raw["Frame Number"].astype(int),
        "Video Time (s)": raw["Video Time (s)"],
        "Elapsed Experiment Time (s)": raw["Video Time (s)"] - start_frame / fps,
        "LED Detected": raw["LED Detected"].astype(bool),
        "Marker 1 X (mm)": marker_mm[:, 0, 0],
        "Marker 1 Y (mm)": marker_mm[:, 0, 1],
        "Marker 2 X (mm)": marker_mm[:, 1, 0],
        "Marker 2 Y (mm)": marker_mm[:, 1, 1],
        "Marker 3 X (mm)": marker_mm[:, 2, 0],
        "Marker 3 Y (mm)": marker_mm[:, 2, 1],
        "Number of Markers Detected": raw["Number of Markers Detected"].astype(int),
        "Interpolated Frame": interpolated_any,
    })
    x_columns = [f"Marker {i} X (mm)" for i in range(1, 4)]
    y_columns = [f"Marker {i} Y (mm)" for i in range(1, 4)]
    output["Object Centroid X (mm)"] = output[x_columns].mean(axis=1, skipna=False)
    output["Object Centroid Y (mm)"] = output[y_columns].mean(axis=1, skipna=False)
    output["Pitch Angle (deg)"] = pitch_from_markers(output)
    smooth_x = output["Object Centroid X (mm)"].rolling(
        SMOOTHING_WINDOW_FRAMES, center=True, min_periods=1
    ).mean()
    # Travel is to image-left, so decreasing global X is positive forward.
    output["Forward Velocity (mm/s)"] = -np.gradient(smooth_x.to_numpy(), 1 / fps)
    crossed = np.zeros(count, dtype=bool)
    if finish_frame is not None:
        crossed[finish_frame:] = True
    output["Finish Line Crossed"] = crossed
    output["Finish Crossing Frame (interpolated)"] = crossing_frame_float
    return output[FRAME_COLUMNS + ["Finish Crossing Frame (interpolated)"]], marker_px_filled


def path_distances(values: np.ndarray) -> tuple[float, float, float]:
    differences = np.diff(values)
    valid = np.isfinite(differences)
    forward = float(np.sum(differences[valid & (differences > MOVEMENT_NOISE_FLOOR_MM)]))
    backward_steps = -differences[valid & (differences < -MOVEMENT_NOISE_FLOOR_MM)]
    backward = float(np.sum(backward_steps))
    maximum_backward = float(np.max(backward_steps)) if len(backward_steps) else 0.0
    return forward, backward, maximum_backward


def cycle_table(
    test_id: str,
    frames: pd.DataFrame,
    blink_frames: list[int],
    start_frame: int,
    finish_frame: int | None,
    fps: float,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    end_limit = finish_frame if finish_frame is not None else len(frames) - 1
    boundaries = [frame for frame in blink_frames if start_frame <= frame <= end_limit]
    rows: list[dict[str, float | int | str]] = []
    rejected: list[dict[str, float | int | str]] = []
    minimum_duration = EXPECTED_CYCLE_DURATION_S * (
        1.0 - COMPLETE_CYCLE_DURATION_TOLERANCE
    )
    maximum_duration = EXPECTED_CYCLE_DURATION_S * (
        1.0 + COMPLETE_CYCLE_DURATION_TOLERANCE
    )
    for candidate_number, (first, last) in enumerate(
        zip(boundaries, boundaries[1:]), 1
    ):
        duration = (last - first) / fps
        if not minimum_duration <= duration <= maximum_duration:
            rejected.append({
                "Test ID": test_id,
                "Candidate Interval Number": candidate_number,
                "Interval Start Time (s)": first / fps,
                "Interval End Time (s)": last / fps,
                "Measured Interval Duration (s)": duration,
                "Expected Cycle Duration (s)": EXPECTED_CYCLE_DURATION_S,
                "Rejection Reason": (
                    f"Duration outside {minimum_duration:.2f}-"
                    f"{maximum_duration:.2f} s complete-cycle range"
                ),
            })
            continue
        segment = frames.iloc[first:last + 1]
        x = segment["Object Centroid X (mm)"].rolling(
            SMOOTHING_WINDOW_FRAMES, center=True, min_periods=1
        ).mean().to_numpy()
        if len(x) < 2 or not np.isfinite(x[[0, -1]]).all():
            continue
        forward_position = x[0] - x
        total_forward, total_backward, maximum_backward = path_distances(forward_position)
        vertical = segment["Object Centroid Y (mm)"].to_numpy(float)
        pitch = segment["Pitch Angle (deg)"].to_numpy(float)
        rows.append({
            "Test ID": test_id,
            "Cycle Number": len(rows) + 1,
            "Cycle Start Time (s)": first / fps,
            "Cycle End Time (s)": last / fps,
            "Measured Cycle Duration (s)": duration,
            "Centroid Start X (mm)": x[0],
            "Centroid End X (mm)": x[-1],
            "Net Advance (mm)": forward_position[-1],
            "Total Forward Movement (mm)": total_forward,
            "Total Backward Movement (mm)": total_backward,
            "Maximum Backward Excursion (mm)": maximum_backward,
            "Vertical Peak-to-Peak (mm)": float(np.nanmax(vertical) - np.nanmin(vertical)),
            "Mean Pitch Angle (deg)": float(np.nanmean(pitch)),
            "Pitch Peak-to-Peak (deg)": float(np.nanmax(pitch) - np.nanmin(pitch)),
        })
    return (
        pd.DataFrame(rows, columns=CYCLE_COLUMNS),
        pd.DataFrame(rejected, columns=REJECTED_CYCLE_COLUMNS),
    )


def run_summary(
    manifest_row: pd.Series,
    frames: pd.DataFrame,
    cycles: pd.DataFrame,
    fps: float,
    start_frame: int,
    finish_frame: int | None,
    crossing_frame_float: float,
    notes: list[str],
) -> dict[str, object]:
    end_frame = finish_frame if finish_frame is not None else len(frames) - 1
    segment = frames.iloc[start_frame:end_frame + 1]
    centroid_x = segment["Object Centroid X (mm)"].rolling(
        SMOOTHING_WINDOW_FRAMES, center=True, min_periods=1
    ).mean().to_numpy(float)
    valid_x = centroid_x[np.isfinite(centroid_x)]
    start_x = float(np.nanmedian(valid_x[:max(1, min(len(valid_x), round(0.2 * fps)))]))
    end_x = float(valid_x[-1]) if len(valid_x) else math.nan
    net_distance = start_x - end_x
    duration = (
        (crossing_frame_float - start_frame) / fps
        if finish_frame is not None else math.nan
    )
    forward_position = start_x - centroid_x
    _, backward, _ = path_distances(forward_position)
    velocity = segment["Forward Velocity (mm/s)"].to_numpy(float)
    vertical = segment["Object Centroid Y (mm)"].to_numpy(float)
    pitch = segment["Pitch Angle (deg)"].to_numpy(float)
    valid_vertical = vertical[np.isfinite(vertical)]
    valid_pitch = pitch[np.isfinite(pitch)]
    detection_rate = 100 * float(
        np.mean(segment["Number of Markers Detected"].to_numpy() == 3)
    )
    qc = "PASS"
    if finish_frame is None:
        qc = "REVIEW - finish not crossed"
    elif detection_rate < 90:
        qc = "REVIEW - marker detection below 90%"
    if not len(valid_x):
        qc = "FAIL - no valid three-marker positions"
    phase = int(manifest_row["Phase Shift (deg)"])
    return {
        "Randomised Run Order": int(manifest_row["Randomised Run Order"]),
        "Test ID": str(manifest_row["Test ID"]),
        "Video Filename": str(manifest_row["Video Filename"]),
        "Phase Shift Command (deg)": phase,
        "Equivalent Signed Phase (deg)": signed_phase(phase),
        "Repeat Number": int(manifest_row["Repeat Number"]),
        "Start Blink Time (s)": start_frame / fps,
        "Finish Crossing Time (s)": crossing_frame_float / fps if finish_frame is not None else math.nan,
        "Transport Time (s)": duration,
        "Net Travel Distance (mm)": net_distance,
        "Mean Transport Speed (mm/s)": net_distance / duration if duration > 0 else math.nan,
        "Median Transport Speed (mm/s)": float(np.nanmedian(velocity)),
        "Number of Complete Cycles": len(cycles),
        "Mean Net Advance per Cycle (mm)": cycles["Net Advance (mm)"].mean(),
        "SD Net Advance per Cycle (mm)": cycles["Net Advance (mm)"].std(ddof=1),
        "Total Backward Travel (mm)": backward,
        "Backward Travel Percentage (%)": 100 * backward / abs(net_distance) if net_distance else math.nan,
        "Maximum Backward Slip in One Cycle (mm)": cycles["Maximum Backward Excursion (mm)"].max(),
        "Vertical Centroid Peak-to-Peak (mm)": float(np.ptp(valid_vertical)) if len(valid_vertical) else math.nan,
        "Vertical Centroid RMS (mm)": float(np.sqrt(np.mean((valid_vertical - np.mean(valid_vertical)) ** 2))) if len(valid_vertical) else math.nan,
        "Mean Pitch Angle (deg)": float(np.mean(valid_pitch)) if len(valid_pitch) else math.nan,
        "Pitch Peak-to-Peak (deg)": float(np.ptp(valid_pitch)) if len(valid_pitch) else math.nan,
        "Maximum Absolute Pitch Angle (deg)": float(np.max(np.abs(valid_pitch))) if len(valid_pitch) else math.nan,
        "Green Marker Detection Rate (%)": detection_rate,
        "Frames Requiring Interpolation": int(segment["Interpolated Frame"].sum()),
        "Finish Crossing Detected": "Yes" if finish_frame is not None else "No",
        "Quality-Control Status": qc,
        "Notes": "; ".join(notes),
    }


def save_run_plot(test_id: str, frames: pd.DataFrame, start: int, end: int, path: Path) -> None:
    segment = frames.iloc[start:end + 1]
    time = segment["Elapsed Experiment Time (s)"]
    centroid_x = segment["Object Centroid X (mm)"]
    forward = float(centroid_x.dropna().iloc[0]) - centroid_x
    vertical = segment["Object Centroid Y (mm)"]
    vertical = vertical - vertical.mean()
    pitch = segment["Pitch Angle (deg)"]
    fig, axes = plt.subplots(3, 1, figsize=(11, 9), sharex=True)
    axes[0].plot(time, forward, color="#137CBD", linewidth=1.7)
    axes[0].set_ylabel("Forward travel (mm)")
    axes[1].plot(time, vertical, color="#2B9B50", linewidth=1.4)
    axes[1].set_ylabel("Vertical deviation (mm)")
    axes[2].plot(time, pitch, color="#D46A1F", linewidth=1.4)
    axes[2].set_ylabel("Pitch (deg)")
    axes[2].set_xlabel("Elapsed time from first LED blink (s)")
    for axis in axes:
        axis.grid(True, alpha=0.28)
    fig.suptitle(test_id)
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def write_overlay(
    video_path: Path,
    output_path: Path,
    calibration: Calibration,
    led_roi: tuple[int, int, int, int],
    frames: pd.DataFrame,
    marker_px: np.ndarray,
    led_threshold: float,
    start_frame: int,
    finish_frame: int | None,
) -> None:
    capture = cv2.VideoCapture(str(video_path))
    fps = float(capture.get(cv2.CAP_PROP_FPS))
    width = int(capture.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(capture.get(cv2.CAP_PROP_FRAME_HEIGHT))
    writer = cv2.VideoWriter(
        str(output_path), cv2.VideoWriter_fourcc(*"mp4v"), fps, (width, height)
    )
    if not writer.isOpened():
        capture.release()
        raise RuntimeError(f"Could not create overlay: {output_path}")
    x, y, roi_width, roi_height = led_roi
    finish_a = tuple(np.round(calibration.finish_points[0]).astype(int))
    finish_b = tuple(np.round(calibration.finish_points[1]).astype(int))
    frame_index = 0
    while frame_index < len(frames):
        ok, raw = capture.read()
        if not ok:
            break
        image = undistort(raw, calibration)
        cv2.rectangle(image, (x, y), (x + roi_width, y + roi_height),
                      (0, 0, 255) if frames.iloc[frame_index]["LED Detected"] else (130, 130, 130), 2)
        cv2.line(image, finish_a, finish_b, (0, 0, 255), 3, cv2.LINE_AA)
        colours = [(0, 255, 0), (0, 220, 255), (255, 170, 0)]
        for marker_id, point in enumerate(marker_px[frame_index]):
            if np.isfinite(point).all():
                centre = tuple(np.round(point).astype(int))
                cv2.circle(image, centre, 13, colours[marker_id], 3, cv2.LINE_AA)
                cv2.putText(image, f"M{marker_id + 1}",
                            (centre[0] + 15, centre[1] - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, colours[marker_id], 2)
        if frame_index < start_frame:
            state, state_colour = "WAITING FOR FIRST RED BLINK", (180, 180, 180)
        elif finish_frame is not None and frame_index >= finish_frame:
            state, state_colour = "FINISH CROSSED", (0, 255, 255)
        else:
            state, state_colour = "TRACKING", (0, 255, 0)
        cv2.rectangle(image, (0, 0), (width, 55), (0, 0, 0), -1)
        cv2.putText(
            image,
            f"{state} | t={frames.iloc[frame_index]['Elapsed Experiment Time (s)']:.2f}s "
            f"| markers={int(frames.iloc[frame_index]['Number of Markers Detected'])}/3",
            (18, 37), cv2.FONT_HERSHEY_SIMPLEX, 0.82, state_colour, 2, cv2.LINE_AA
        )
        writer.write(image)
        frame_index += 1
    capture.release()
    writer.release()


def phase_comparison(summary: pd.DataFrame) -> pd.DataFrame:
    valid = summary[
        summary["Finish Crossing Detected"].eq("Yes")
        & summary["Mean Transport Speed (mm/s)"].notna()
    ]
    rows: list[dict[str, object]] = []
    for phase in EXPECTED_PHASES:
        group = valid[valid["Phase Shift Command (deg)"].eq(phase)]
        mean_time = group["Transport Time (s)"].mean()
        sd_time = group["Transport Time (s)"].std(ddof=1)
        rows.append({
            "Phase Shift Command (deg)": phase,
            "Equivalent Signed Phase (deg)": signed_phase(phase),
            "Valid Repeats": len(group),
            "Mean Transport Time (s)": mean_time,
            "SD Transport Time (s)": sd_time,
            "Coefficient of Variation (%)": 100 * sd_time / mean_time if mean_time else math.nan,
            "Mean Transport Speed (mm/s)": group["Mean Transport Speed (mm/s)"].mean(),
            "SD Transport Speed (mm/s)": group["Mean Transport Speed (mm/s)"].std(ddof=1),
            "Mean Advance per Cycle (mm)": group["Mean Net Advance per Cycle (mm)"].mean(),
            "Mean Total Backward Travel (mm)": group["Total Backward Travel (mm)"].mean(),
            "Mean Backward Travel Percentage (%)": group["Backward Travel Percentage (%)"].mean(),
            "Mean Vertical Peak-to-Peak (mm)": group["Vertical Centroid Peak-to-Peak (mm)"].mean(),
            "Mean Pitch Peak-to-Peak (deg)": group["Pitch Peak-to-Peak (deg)"].mean(),
        })
    result = pd.DataFrame(rows)
    ranks = result["Mean Transport Speed (mm/s)"].rank(
        ascending=False, method="min", na_option="bottom"
    )
    result["Performance Rank"] = ranks.where(result["Valid Repeats"] > 0).astype("Int64")
    return result


def save_phase_plot(comparison: pd.DataFrame, path: Path) -> None:
    x = np.arange(len(comparison))
    labels = [str(int(value)) for value in comparison["Phase Shift Command (deg)"]]
    fig, axes = plt.subplots(2, 2, figsize=(12, 8.5))
    plots = [
        ("Mean Transport Speed (mm/s)", "SD Transport Speed (mm/s)", "Speed (mm/s)"),
        ("Mean Total Backward Travel (mm)", None, "Backward travel (mm)"),
        ("Mean Vertical Peak-to-Peak (mm)", None, "Vertical P-P (mm)"),
        ("Mean Pitch Peak-to-Peak (deg)", None, "Pitch P-P (deg)"),
    ]
    for axis, (column, error_column, ylabel) in zip(axes.flat, plots):
        values = comparison[column].to_numpy(float)
        errors = comparison[error_column].to_numpy(float) if error_column else None
        axis.bar(x, values, yerr=errors, capsize=4, color="#3B82B8")
        axis.set_xticks(x, labels)
        axis.set_xlabel("Commanded phase shift (deg)")
        axis.set_ylabel(ylabel)
        axis.grid(axis="y", alpha=0.25)
    fig.suptitle("Phase experiment comparison (mean of valid repeats)")
    fig.tight_layout()
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def metric_definitions() -> pd.DataFrame:
    return pd.DataFrame([
        ("Transport time", "First detected red-LED rising edge to interpolated finish-line crossing."),
        ("Mean transport speed", "Net centroid travel divided by transport time."),
        ("Complete cycle", "Consecutive LED rising edges separated by 2.40-3.60 s (3.0 s expected cycle, +/-20% tolerance)."),
        ("Net advance per cycle", "Decrease in centroid X across an accepted complete cycle; partial start/stop intervals are excluded."),
        ("Backward travel", "Sum of smoothed reverse increments greater than the 0.20 mm noise floor."),
        ("Vertical peak-to-peak", "Maximum minus minimum three-marker centroid height during the run."),
        ("Vertical RMS", "RMS centroid-height deviation about that run's mean height."),
        ("Pitch", "Angle of the line from marker 1 to marker 3; positive Y is reported upward."),
        ("Detection rate", "Percentage of analysed frames in which all three dots were directly detected."),
        ("Interpolated frame", "At least one marker filled across a gap of at most five frames."),
        ("Performance rank", "Rank by mean transport speed; fastest valid phase is rank 1."),
    ], columns=["Metric", "Definition"])


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[1].height = 42
        for column_cells in sheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:200]]
            width = min(42, max(10, max(len(value) for value in values) + 2))
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    workbook.save(path)


def write_workbook(
    summary: pd.DataFrame,
    comparison: pd.DataFrame,
    cycles: pd.DataFrame,
    rejected_cycles: pd.DataFrame,
    frames: pd.DataFrame,
    manifest: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(WORKBOOK_PATH, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Run Summary", index=False)
        comparison.to_excel(writer, sheet_name="Phase Comparison", index=False)
        cycles.to_excel(writer, sheet_name="Cycle Data", index=False)
        rejected_cycles.to_excel(writer, sheet_name="Rejected Cycles", index=False)
        frames.to_excel(writer, sheet_name="Frame Data", index=False)
        manifest.to_excel(writer, sheet_name="Experiment Log", index=False)
        metric_definitions().to_excel(writer, sheet_name="Metric Definitions", index=False)
    style_workbook(WORKBOOK_PATH)


def main() -> None:
    args = parse_args()
    calibration = load_calibration(CALIBRATION_NPZ)
    videos = discover_videos()
    manifest = create_or_update_manifest(videos)
    validate_manifest(manifest)
    included = manifest[
        pd.to_numeric(manifest["Include"], errors="coerce").fillna(0) != 0
    ].copy()
    included = included.sort_values("Randomised Run Order")
    first_video = EXPERIMENT_VIDEO_DIR / str(included.iloc[0]["Video Filename"])
    led_roi = select_or_load_led_roi(first_video, calibration, args.reset_led_roi)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    summaries: list[dict[str, object]] = []
    all_cycles: list[pd.DataFrame] = []
    all_rejected_cycles: list[pd.DataFrame] = []
    all_frames: list[pd.DataFrame] = []
    for position, (_, manifest_row) in enumerate(included.iterrows(), 1):
        video_path = EXPERIMENT_VIDEO_DIR / str(manifest_row["Video Filename"])
        test_id = str(manifest_row["Test ID"])
        run_folder = OUTPUT_DIR / test_id
        run_folder.mkdir(parents=True, exist_ok=True)
        print(f"\n[{position}/{len(included)}] Analysing {video_path.name}")
        notes: list[str] = []
        try:
            raw, metadata, marker_px = track_raw_video(video_path, calibration, led_roi)
            fps = float(metadata["fps"])
            _, blink_frames, led_threshold = detect_led_blinks(raw, fps)
            if not blink_frames:
                raise RuntimeError("No valid red LED blink was detected.")
            start_frame = blink_frames[0]
            leading_px = marker_px[:, 0, :]
            finish_frame, crossing_float = first_finish_crossing(
                leading_px, start_frame, calibration.finish_line
            )
            if len(blink_frames) < 2:
                notes.append("Only one LED rising edge detected; cycle metrics unavailable")
            if finish_frame is None:
                notes.append("Leading marker did not cross the calibrated finish line")
                crossing_float = math.nan

            frames, filled_marker_px = prepare_frame_data(
                test_id, raw, marker_px, calibration, fps, start_frame,
                finish_frame, crossing_float
            )
            cycles, rejected_cycles = cycle_table(
                test_id, frames, blink_frames, start_frame, finish_frame, fps
            )
            if not rejected_cycles.empty:
                notes.append(
                    f"{len(rejected_cycles)} incomplete LED interval(s) excluded "
                    "from cycle metrics"
                )
            summary = run_summary(
                manifest_row, frames, cycles, fps, start_frame,
                finish_frame, crossing_float, notes
            )
            analysis_end = finish_frame if finish_frame is not None else len(frames) - 1
            frames.iloc[start_frame:analysis_end + 1].to_csv(
                run_folder / "frame_tracking.csv", index=False
            )
            cycles.to_csv(run_folder / "cycle_metrics.csv", index=False)
            rejected_cycles.to_csv(
                run_folder / "rejected_cycle_intervals.csv", index=False
            )
            pd.DataFrame([summary]).to_csv(run_folder / "run_summary.csv", index=False)
            led_audit = raw[["Frame Number", "Video Time (s)", "LED Score",
                             "LED Score Smoothed", "LED Detected"]].copy()
            led_audit["Detection Threshold"] = led_threshold
            led_audit.to_csv(run_folder / "led_detection.csv", index=False)
            save_run_plot(
                test_id, frames, start_frame, analysis_end,
                run_folder / "motion_diagnostics.png"
            )
            if WRITE_ANNOTATED_VIDEOS and not args.no_overlay:
                write_overlay(
                    video_path, run_folder / "tracking_overlay.mp4", calibration,
                    led_roi, frames, filled_marker_px, led_threshold,
                    start_frame, finish_frame
                )
            summaries.append(summary)
            all_cycles.append(cycles)
            all_rejected_cycles.append(rejected_cycles)
            all_frames.append(frames.iloc[start_frame:analysis_end + 1])
            print(
                f"  {summary['Quality-Control Status']} | "
                f"time={summary['Transport Time (s)']} s | "
                f"3-dot detection={summary['Green Marker Detection Rate (%)']:.1f}%"
            )
        except Exception as error:
            print(f"  FAILED: {error}")
            phase = int(manifest_row["Phase Shift (deg)"])
            failed = {column: math.nan for column in RUN_SUMMARY_COLUMNS}
            failed.update({
                "Randomised Run Order": int(manifest_row["Randomised Run Order"]),
                "Test ID": test_id, "Video Filename": video_path.name,
                "Phase Shift Command (deg)": phase,
                "Equivalent Signed Phase (deg)": signed_phase(phase),
                "Repeat Number": int(manifest_row["Repeat Number"]),
                "Finish Crossing Detected": "No",
                "Quality-Control Status": "FAIL - processing error",
                "Notes": str(error),
            })
            summaries.append(failed)

    summary_df = pd.DataFrame(summaries, columns=RUN_SUMMARY_COLUMNS).sort_values(
        "Randomised Run Order"
    )
    cycles_df = (
        pd.concat(all_cycles, ignore_index=True)
        if all_cycles else pd.DataFrame(columns=CYCLE_COLUMNS)
    )
    rejected_cycles_df = (
        pd.concat(all_rejected_cycles, ignore_index=True)
        if all_rejected_cycles
        else pd.DataFrame(columns=REJECTED_CYCLE_COLUMNS)
    )
    frames_df = (
        pd.concat(all_frames, ignore_index=True)
        if all_frames else pd.DataFrame(columns=FRAME_COLUMNS)
    )
    comparison_df = phase_comparison(summary_df)
    write_workbook(
        summary_df,
        comparison_df,
        cycles_df,
        rejected_cycles_df,
        frames_df,
        manifest,
    )
    save_phase_plot(comparison_df, OUTPUT_DIR / "phase_comparison.png")
    summary_df.to_csv(OUTPUT_DIR / "run_summary.csv", index=False)
    comparison_df.to_csv(OUTPUT_DIR / "phase_comparison.csv", index=False)

    print("\nBatch analysis complete.")
    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Comparison plot: {OUTPUT_DIR / 'phase_comparison.png'}")
    review_count = int(summary_df["Quality-Control Status"].astype(str).str.contains(
        "REVIEW|FAIL", regex=True
    ).sum())
    print(f"Runs requiring review: {review_count} of {len(summary_df)}")
    print(f"Incomplete cycle intervals excluded: {len(rejected_cycles_df)}")


if __name__ == "__main__":
    main()
